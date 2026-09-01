"""Write-through mirror of each user's own queue into their own tracker.xlsx.

One file per user under TRACKER_DIR - never a single shared file. This used
to be one shared tracker.xlsx for every user (a real cross-user data leak:
every user's subject/company/role/hr_email/status ended up in the same
workbook, keyed only by gmail_message_id, with no user_id column at all).
Not reachable through any HTTP route today (verified: no route reads,
downloads, or serves this file), but the commingling itself was real and
inconsistent with every other per-user-scoped table/resource in this app -
fixed for defense-in-depth and consistency, not because it was live-exploitable."""
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook

from config import TRACKER_DIR

HEADERS = [
    "date_processed",
    "gmail_message_id",
    "subject",
    "company",
    "role",
    "hr_email_used",
    "status",
    "resume_filename",
    "sent_at",
]


def _tracker_file(user_id: int):
    return TRACKER_DIR / f"{user_id}.xlsx"


def _load_or_create(user_id: int):
    tracker_file = _tracker_file(user_id)
    if tracker_file.exists():
        wb = load_workbook(tracker_file)
        ws = wb.active
    else:
        tracker_file.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
    return wb, ws


def upsert_tracker_row(user_id: int, queue_row: dict) -> None:
    wb, ws = _load_or_create(user_id)

    message_id_col = HEADERS.index("gmail_message_id") + 1
    target_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=message_id_col).value == queue_row["gmail_message_id"]:
            target_row = r
            break

    values = [
        datetime.now(timezone.utc).isoformat(),
        queue_row.get("gmail_message_id"),
        queue_row.get("subject"),
        queue_row.get("company"),
        queue_row.get("role"),
        queue_row.get("hr_email"),
        queue_row.get("status"),
        queue_row.get("resume_filename"),
        queue_row.get("sent_at"),
    ]

    if target_row is None:
        ws.append(values)
    else:
        for col, value in enumerate(values, start=1):
            ws.cell(row=target_row, column=col, value=value)

    wb.save(_tracker_file(user_id))
