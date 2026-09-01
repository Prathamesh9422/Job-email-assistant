"""Multi-user data isolation audit — automated cross-user tests.

Covers the 10 areas from the security audit: queue isolation, sent-history
isolation, Gmail credential isolation, resume isolation, Chrome extension
token isolation, scheduler/fetch isolation, per-action ownership checks,
session/logout switching, IDOR via a foreign row id, and OAuth state/session
protection.

No real Gmail API call or real email send happens anywhere in this file —
every Gmail-touching function is mocked. Uses a throwaway SQLite file per
test (never the real dev/prod database).
"""
from unittest.mock import MagicMock, patch

import pytest
from sqlalchemy import create_engine
from starlette.testclient import TestClient

import db


@pytest.fixture()
def ctx(tmp_path, monkeypatch):
    """Two independent users, a shared TestClient, and a per-user session
    switcher — all against a throwaway SQLite file and a throwaway resume folder."""
    engine = create_engine(f"sqlite:///{tmp_path / 'test.db'}", future=True)
    monkeypatch.setattr(db, "_engine", engine)
    db.init_db()

    import app
    import auth
    import excel_tracker
    import fetch_job

    # Keep resume/tracker files out of the real project folder entirely.
    monkeypatch.setattr(app, "RESUMES_DIR", tmp_path / "resumes")
    monkeypatch.setattr(excel_tracker, "TRACKER_DIR", tmp_path / "trackers")

    user_a = db.upsert_user("alice@test.com", "sub-alice", "refresh-token-alice")
    user_b = db.upsert_user("bob@test.com", "sub-bob", "refresh-token-bob")

    client = TestClient(app.app)

    def as_user(user):
        app.app.dependency_overrides[auth.require_user] = lambda: user

    yield {
        "app": app,
        "auth": auth,
        "excel_tracker": excel_tracker,
        "fetch_job": fetch_job,
        "client": client,
        "a": user_a,
        "b": user_b,
        "as_user": as_user,
    }
    app.app.dependency_overrides.clear()


def _insert_row(user_id: int, gmail_message_id: str, **overrides) -> int:
    row = {
        "gmail_message_id": gmail_message_id,
        "received_at": "2026-01-01T00:00:00+00:00",
        "subject": "Test job",
        "company": "Acme Corp",
        "role": "Engineer",
        "hr_email": "hr@acme.example.com",
        "status": "ready",
    }
    row.update(overrides)
    with db.get_conn() as conn:
        row_id = db.insert_queue_row(conn, user_id, row)
    assert row_id is not None
    return row_id


# --- 1. Queue/job isolation ---


def test_queue_isolation(ctx):
    row_a = _insert_row(ctx["a"]["id"], "msg-a-1")
    _insert_row(ctx["b"]["id"], "msg-b-1")

    ctx["as_user"](ctx["a"])
    resp = ctx["client"].get("/api/queue?status=ready")
    ids = [r["id"] for r in resp.json()]
    assert row_a in ids
    assert len(resp.json()) == 1  # only A's row, never B's


# --- 2. Sent-email history isolation ---


def test_sent_history_isolation(ctx):
    row_a = _insert_row(ctx["a"]["id"], "msg-a-2")
    db.insert_sent_record(ctx["a"]["id"], row_a, "Subject A", "Body A")

    ctx["as_user"](ctx["b"])
    resp = ctx["client"].get(f"/api/queue/{row_a}/sent-records")
    assert resp.status_code == 404  # B doesn't even see that the row exists

    ctx["as_user"](ctx["a"])
    resp = ctx["client"].get(f"/api/queue/{row_a}/sent-records")
    assert resp.status_code == 200
    assert len(resp.json()) == 1
    assert resp.json()[0]["subject"] == "Subject A"


# --- 2b. Excel tracker isolation (ARC-0002's previously-open question) ---


def test_excel_tracker_isolation(ctx):
    """tracker.xlsx used to be one shared file for every user - this proves
    it's now one file per user, and that writing A's row never touches or
    creates B's file."""
    excel_tracker = ctx["excel_tracker"]

    row_a = _insert_row(ctx["a"]["id"], "msg-tracker-a", company="Alice Co")
    row = db.get_queue_row(ctx["a"]["id"], row_a)
    excel_tracker.upsert_tracker_row(ctx["a"]["id"], row)

    tracker_a = excel_tracker._tracker_file(ctx["a"]["id"])
    tracker_b = excel_tracker._tracker_file(ctx["b"]["id"])

    assert tracker_a.exists()
    assert not tracker_b.exists()  # B's file was never created or touched

    from openpyxl import load_workbook

    ws = load_workbook(tracker_a).active
    companies = [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)]
    assert companies == ["Alice Co"]


# --- 3. Gmail credential isolation ---


def test_gmail_credential_isolation(ctx):
    auth = ctx["auth"]
    with patch.object(auth, "GoogleRequest"), patch(
        "google.oauth2.credentials.Credentials.refresh", return_value=None
    ):
        cred_a = auth.get_credential_for_user(ctx["a"]["id"])
        cred_b = auth.get_credential_for_user(ctx["b"]["id"])

    assert cred_a.refresh_token == "refresh-token-alice"
    assert cred_b.refresh_token == "refresh-token-bob"
    assert cred_a.refresh_token != cred_b.refresh_token


# --- 4. Resume isolation (the confirmed vulnerability) ---


def test_resume_isolation(ctx):
    app = ctx["app"]

    dir_a = app._user_resume_dir(ctx["a"]["id"])
    dir_b = app._user_resume_dir(ctx["b"]["id"])
    dir_a.mkdir(parents=True)
    dir_b.mkdir(parents=True)
    (dir_a / "alice_resume.pdf").write_bytes(b"%PDF-1.4 alice")
    (dir_b / "bob_resume.pdf").write_bytes(b"%PDF-1.4 bob")

    ctx["as_user"](ctx["a"])
    resp = ctx["client"].get("/api/resume")
    assert resp.json()["resume_filename"] == "alice_resume.pdf"

    ctx["as_user"](ctx["b"])
    resp = ctx["client"].get("/api/resume")
    assert resp.json()["resume_filename"] == "bob_resume.pdf"


def test_resume_upload_is_per_user(ctx):
    ctx["as_user"](ctx["a"])
    resp = ctx["client"].post(
        "/api/resume", files={"file": ("alice.pdf", b"%PDF-1.4", "application/pdf")}
    )
    assert resp.status_code == 200

    ctx["as_user"](ctx["b"])
    resp = ctx["client"].get("/api/resume")
    assert resp.json()["resume_filename"] is None  # A's upload never visible to B


# --- 5. Chrome extension token isolation ---


def test_extension_token_isolation(ctx):
    auth = ctx["auth"]
    token_a = auth.generate_api_token(ctx["a"]["id"])

    resolved = db.get_user_by_api_token_hash(auth._hash_api_token(token_a))
    assert resolved["id"] == ctx["a"]["id"]
    assert resolved["id"] != ctx["b"]["id"]

    resp = ctx["client"].post(
        "/api/queue/ingest-scraped",
        json={"jobs": [{"company": "X", "role": "Y", "applied_date": "today"}]},
        headers={"Authorization": f"Bearer {token_a}"},
    )
    assert resp.status_code == 200

    # The row that landed must belong to A, not B.
    ctx["as_user"](ctx["b"])
    assert ctx["client"].get("/api/queue?status=needs_info,ready").json() == []
    ctx["as_user"](ctx["a"])
    rows = ctx["client"].get("/api/queue?status=needs_info,ready").json()
    assert any(r["company"] == "X" for r in rows)


def test_extension_token_expires(ctx):
    """Hardening check (not one of the 10 core areas): a token older than
    auth.API_TOKEN_TTL_DAYS must stop working."""
    auth = ctx["auth"]
    token = auth.generate_api_token(ctx["a"]["id"])

    with db.get_conn() as conn:
        from sqlalchemy import text

        conn.execute(
            text("UPDATE users SET api_token_created_at = :ts WHERE id = :id"),
            {"ts": "2000-01-01T00:00:00+00:00", "id": ctx["a"]["id"]},
        )
        conn.commit()

    resp = ctx["client"].post(
        "/api/queue/ingest-scraped",
        json={"jobs": []},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert resp.status_code == 401


# --- 6. Scheduler/fetch isolation ---


def test_scheduler_fetch_isolation(ctx):
    fetch_job = ctx["fetch_job"]
    auth = ctx["auth"]

    def fake_credentials_for(user_id):
        creds = MagicMock()
        creds.owner_user_id = user_id  # so the mocked Gmail calls below can tell users apart
        return creds

    def fake_search(credentials, query):
        # Real Gmail message ids are globally unique (Gmail's own storage
        # assigns them, not per-mailbox) - that's what the `queue` table's
        # global UNIQUE constraint on gmail_message_id relies on for dedup.
        # Use a distinct id per user so this test doesn't accidentally
        # collide the two users' rows against that constraint.
        return [f"msg-user-{credentials.owner_user_id}"]

    def fake_get_message(credentials, msg_id):
        return {
            "id": msg_id,
            "threadId": "t1",
            "payload": {"headers": [{"name": "Subject", "value": "Job | Engineer in Pune"}]},
        }

    with patch.object(auth, "get_credential_for_user", side_effect=fake_credentials_for), patch(
        "gmail_client.search_messages", side_effect=fake_search
    ), patch("gmail_client.get_message", side_effect=fake_get_message):
        result_a = fetch_job.run_fetch_for_user(ctx["a"])
        result_b = fetch_job.run_fetch_for_user(ctx["b"])

    assert result_a["status"] == "success"
    assert result_b["status"] == "success"

    rows_a = db.list_queue(ctx["a"]["id"])
    rows_b = db.list_queue(ctx["b"]["id"])
    assert len(rows_a) == 1
    assert len(rows_b) == 1
    assert rows_a[0]["id"] != rows_b[0]["id"]


# --- 7 & 9. Per-action ownership checks + IDOR via foreign row id ---


@pytest.mark.parametrize(
    "method,path_suffix,body",
    [
        ("get", "", None),
        ("patch", "", {"hr_email": "attacker@evil.example.com"}),
        ("post", "/scrape", None),
        ("post", "/preview", {"template": "cold_outreach"}),
        ("post", "/finalize", {"subject": "x", "body": "y"}),
        ("post", "/send", None),
        ("post", "/skip", None),
        ("post", "/advance", {"event": "reject"}),
        ("get", "/sent-records", None),
    ],
)
def test_ownership_checks_all_actions(ctx, method, path_suffix, body):
    row_a = _insert_row(ctx["a"]["id"], f"msg-idor-{path_suffix or 'get'}")

    ctx["as_user"](ctx["b"])
    call = getattr(ctx["client"], method)
    path = f"/api/queue/{row_a}{path_suffix}"
    resp = call(path, json=body) if body is not None else call(path)

    assert resp.status_code == 404, f"{method.upper()} {path} leaked or mutated A's row for B"

    # Confirm the row was not silently mutated either (only meaningful for
    # patch/finalize/advance/skip, but cheap to check unconditionally).
    ctx["as_user"](ctx["a"])
    row_after = ctx["client"].get(f"/api/queue/{row_a}").json()
    assert row_after["hr_email"] != "attacker@evil.example.com"


# --- 8. Session/logout/login switching ---


class _FakeRequest:
    """Starlette's Request.session is just request.scope['session'], a plain
    dict populated by SessionMiddleware from the signed cookie. Testing
    auth.logout()/require_user() against a bare object with that one
    attribute exercises their real logic without needing to forge a signed
    session cookie or fight the test client's cookie jar."""

    def __init__(self, session: dict):
        self.session = session


def test_logout_clears_session(ctx):
    auth = ctx["auth"]
    request = _FakeRequest({"user_id": ctx["a"]["id"], "oauth_state": "leftover"})

    auth.logout(request)

    assert request.session == {}


def test_login_switch_leaves_no_residue_of_previous_user(ctx):
    """Logout then a fresh sign-in must not carry over the old user_id."""
    auth = ctx["auth"]
    request = _FakeRequest({"user_id": ctx["a"]["id"]})

    auth.logout(request)
    assert "user_id" not in request.session

    request.session["user_id"] = ctx["b"]["id"]
    resolved = auth.require_user(request)
    assert resolved["id"] == ctx["b"]["id"]
    assert resolved["id"] != ctx["a"]["id"]


def test_require_user_rejects_cleared_session(ctx):
    auth = ctx["auth"]
    request = _FakeRequest({})

    with pytest.raises(Exception) as exc_info:
        auth.require_user(request)
    assert getattr(exc_info.value, "status_code", None) == 401


# --- 10. OAuth state/session protection ---


def test_oauth_state_mismatch_rejected(ctx):
    app = ctx["app"]
    auth = ctx["auth"]
    app.app.dependency_overrides.pop(auth.require_user, None)

    with TestClient(app.app) as c:
        # No oauth_state ever stored in this session, so any state in the
        # callback query string is automatically a mismatch.
        resp = c.get("/auth/callback?state=attacker-supplied-state&code=whatever", follow_redirects=False)
        assert resp.status_code == 400
